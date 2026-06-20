from pathlib import Path
import json
import re
import sys
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

ROOT = Path(__file__).resolve().parents[1]
DOWNLOADS = Path("D:/Download")
ACC_PATH = DOWNLOADS / "ACC  Exprogram Test Apr20.xlsx"
OUT = ROOT / "src" / "data" / "excelSeed.js"


def clean(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def number(value):
    if value is None or value == "":
        return 0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0


def extract_accounts(workbook):
    ws = workbook["Balance Be"]
    accounts = []
    for row in ws.iter_rows(min_row=5, max_row=80, values_only=True):
        code = clean(row[2])
        name_lao = clean(row[3])
        opening_dr = number(row[4])
        opening_cr = number(row[5])
        move_dr = number(row[6])
        move_cr = number(row[7])
        close_dr = number(row[8])
        close_cr = number(row[9])
        if not code or not code.isdigit() or len(code) < 4:
            continue
        name_lao = LAO_NAMES.get(code, name_lao if not looks_mojibake(name_lao) else "")
        balance = max(opening_dr + move_dr - opening_cr - move_cr, opening_cr + move_cr - opening_dr - move_dr)
        accounts.append(
            {
                "code": code,
                "nameLao": name_lao,
                "nameEn": ACCOUNT_NAMES_EN.get(code, "Mapped account"),
                "type": account_type(code),
                "balance": round(balance),
                "debit": round(close_dr),
                "credit": round(close_cr),
            }
        )
    return accounts[:14]


def extract_journal(workbook):
    ws = workbook["GL"]
    entries = []
    for row in ws.iter_rows(min_row=5, max_row=45, values_only=True):
        entry_no = clean(row[0])
        description = clean(row[3])
        debit_account = clean(row[4])
        credit_account = clean(row[5])
        debit_amount = number(row[6])
        credit_amount = number(row[7])
        if not entry_no:
            continue
        if description and not debit_account and not credit_account:
            current = {
                "id": f"JV-2025-{len(entries) + 1:03d}",
                "description": description,
                "lines": [],
            }
            entries.append(current)
            continue
        if not entries or (not debit_account and not credit_account):
            continue
        entries[-1]["lines"].append(
            {
                "account": debit_account or credit_account,
                "side": "debit" if debit_account else "credit",
                "amount": round(debit_amount or credit_amount),
            }
        )
    return [entry for entry in entries if entry["lines"]][:6]


def account_type(code):
    first = code[:1]
    return {
        "1": "Asset",
        "2": "Liability",
        "3": "Equity",
        "4": "Equity",
        "5": "Expense",
        "6": "Expense",
        "7": "Revenue",
    }.get(first, "Other")


def looks_mojibake(value):
    return any(marker in value for marker in ("à", "â", "€", "º", "»"))


LAO_NAMES = {
    "1011": "ເງິນສົດເປັນກີບ",
    "1012": "ເງິນສົດເປັນເງິນຕ່າງປະເທດ",
    "1013": "ເງິນສົດຍ່ອຍ",
    "1014": "ເງິນສົດກຳລັງນຳຝາກ",
    "1017": "ເງິນສົດລ່ວງໜ້າ",
    "1021": "ເງິນຝາກທະນາຄານເປັນກີບ",
    "1022": "ເງິນຝາກທະນາຄານເປັນເງິນຕ່າງປະເທດ",
    "1023": "ເງິນຝາກຄັງເງິນແຫ່ງຊາດ",
    "1024": "ເງິນຝາກສະຖາບັນການເງິນອື່ນໆ",
    "1028": "ດອກເບ້ຍຕ້ອງຮັບ",
    "1111": "ເງິນກູ້ຢືມໄລຍະສັ້ນ",
    "1118": "ດອກເບ້ຍຄ້າງຮັບ",
    "1121": "ລູກໜີ້ການຄ້າ",
    "1211": "ສິນຄ້າຄົງເຫຼືອ",
}


ACCOUNT_NAMES_EN = {
    "1011": "Cash in LAK",
    "1012": "Cash in foreign currency",
    "1013": "Petty cash",
    "1014": "Cash in transit",
    "1017": "Cash advance",
    "1021": "Bank deposits in LAK",
    "1022": "Bank deposits in foreign currency",
    "1023": "Treasury deposits",
    "1024": "Other financial deposits",
    "1028": "Interest receivable",
    "1111": "Short-term loans receivable",
    "1118": "Accrued interest receivable",
    "1121": "Trade receivables",
    "1211": "Inventory",
}


def summarize_workbook(workbook):
    return [
        {
            "sheet": ws.title,
            "dimension": ws.calculate_dimension(),
            "values": sum(1 for row in ws.iter_rows() for cell in row if cell.value is not None),
        }
        for ws in workbook.worksheets
    ]


def main():
    workbook = load_workbook(ACC_PATH, read_only=True, data_only=True)
    data = {
        "sourceFile": str(ACC_PATH),
        "accounts": extract_accounts(workbook),
        "journalEntries": extract_journal(workbook),
        "workbookSummary": summarize_workbook(workbook),
    }
    OUT.write_text(
        "export const excelSeed = "
        + json.dumps(data, ensure_ascii=False, indent=2)
        + ";\n",
        encoding="utf-8",
    )
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
